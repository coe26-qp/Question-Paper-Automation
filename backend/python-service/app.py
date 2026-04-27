# from flask import Flask, request, jsonify
# from openpyxl import load_workbook
# import base64
# import traceback

# app = Flask(__name__)

# # safe anchor reader
# def get_anchor(img):
#     try:
#         if hasattr(img.anchor, "_from"):
#             return img.anchor._from.row + 1, img.anchor._from.col + 1
#         elif hasattr(img.anchor, "from_"):
#             return img.anchor.from_.row + 1, img.anchor.from_.col + 1
#         else:
#             return None, None
#     except:
#         return None, None

# @app.route("/")
# def home():
#     return "Flask API is running!"


# @app.route("/extract", methods=["POST"])
# def extract_questions():
#     try:
#         # 1️⃣ Check file
#         if "file" not in request.files:
#             return jsonify({"error": "No file uploaded"}), 400

#         file = request.files["file"]

#         # 2️⃣ Load Excel
#         wb = load_workbook(file)
#         ws = wb.active

#         results = []

#         # 3️⃣ Loop rows (skip header)
#         for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):

#             if not any(cell.value for cell in row):
#                 continue

#             unit = row[0].value
#             co = row[1].value
#             unit_title = row[2].value
#             qtype = row[3].value
#             part = row[4].value
#             question_text = row[5].value
#             marks = row[6].value
#             bt = row[7].value
#             level = row[8].value

#             # normalize TYPE (important for Mongo enum)
#             if qtype:
#                 qtype = str(qtype).strip().upper()

#             # ---------- IMAGE DETECTION ----------
#             question_image = None

#             for img in ws._images:
#                 img_row, img_col = get_anchor(img)

#                 # Question column = 6
#                 if img_row == row_index and img_col == 6:
#                     img_bytes = img._data()
#                     question_image = base64.b64encode(img_bytes).decode("utf-8")
#                     break

#             # ---------- STORE BOTH ----------
#             results.append({
#                 "unit": unit,
#                 "co": co,
#                 "unit_title": unit_title,
#                 "type": qtype,
#                 "part": part,
#                 "marks": marks,
#                 "bt": bt,
#                 "level": level,
#                 "questionText": question_text if question_text else None,
#                 "questionImage": question_image
#             })

#         # ⭐⭐⭐ THE MOST IMPORTANT LINE ⭐⭐⭐
#         return jsonify(results)

#     except Exception as e:
#         print("PYTHON EXTRACTION ERROR:\n", traceback.format_exc())
#         return jsonify({
#             "error": "Extraction failed",
#             "details": str(e)
#         }), 500


# # if __name__ == "__main__":
# #     app.run(port=5001)


# if __name__ == "__main__":
#     import os
#     port = int(os.environ.get("PORT", 5001))
#     app.run(host="0.0.0.0", port=port)




# from flask import Flask, request, jsonify
# from openpyxl import load_workbook
# import base64
# import traceback
# from flask_cors import CORS
# import zipfile
# from io import BytesIO

# app = Flask(__name__)
# CORS(app)

# # -----------------------------
# # Safe anchor reader
# # -----------------------------
# def get_anchor(img):
#     try:
#         if hasattr(img.anchor, "_from"):
#             return img.anchor._from.row + 1, img.anchor._from.col + 1
#         elif hasattr(img.anchor, "from_"):
#             return img.anchor.from_.row + 1, img.anchor.from_.col + 1
#         else:
#             return None, None
#     except:
#         return None, None

# # -----------------------------
# # Extract images using ZIP (cloud-safe)
# # -----------------------------
# def extract_images_from_excel(file_bytes):
#     image_list = []
#     try:
#         with zipfile.ZipFile(BytesIO(file_bytes)) as z:
#             for name in z.namelist():
#                 if name.startswith("xl/media/"):
#                     img_data = z.read(name)
#                     base64_img = base64.b64encode(img_data).decode("utf-8")
#                     image_list.append(base64_img)
#     except Exception as e:
#         print("ZIP extraction error:", str(e))
#     return image_list

# # -----------------------------
# # Health check
# # -----------------------------
# @app.route("/")
# def home():
#     return "Flask API is running!"

# # -----------------------------
# # Main API
# # -----------------------------
# @app.route("/extract", methods=["POST"])
# def extract_questions():
#     try:
#         if "file" not in request.files:
#             return jsonify({"error": "No file uploaded"}), 400

#         file = request.files["file"]

#         # Read file as bytes (for ZIP fallback)
#         file_bytes = file.read()

#         # Reset pointer for openpyxl
#         file_stream = BytesIO(file_bytes)

#         # Load Excel
#         wb = load_workbook(file_stream, data_only=True)
#         ws = wb.active

#         results = []

#         # Try openpyxl images
#         images = getattr(ws, "_images", [])
#         print("openpyxl images found:", len(images))

#         # Fallback: extract images via ZIP
#         zip_images = extract_images_from_excel(file_bytes)
#         print("ZIP images found:", len(zip_images))

#         zip_index = 0

#         # Loop rows
#         for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):

#             if not any(cell.value for cell in row):
#                 continue

#             unit = row[0].value
#             co = row[1].value
#             unit_title = row[2].value
#             qtype = row[3].value
#             part = row[4].value
#             question_text = row[5].value
#             marks = row[6].value
#             bt = row[7].value
#             level = row[8].value

#             if qtype:
#                 qtype = str(qtype).strip().upper()

#             # -----------------------------
#             # IMAGE DETECTION
#             # -----------------------------
#             question_image = None

#             # 1️⃣ Try openpyxl (precise mapping)
#             for img in images:
#                 try:
#                     img_row, img_col = get_anchor(img)

#                     if img_row == row_index and img_col == 6:
#                         img_bytes = img._data()
#                         question_image = base64.b64encode(img_bytes).decode("utf-8")
#                         break
#                 except Exception as e:
#                     print("Image error:", str(e))

#             # 2️⃣ Fallback (ZIP extraction)
#             if not question_image and zip_index < len(zip_images):
#                 question_image = zip_images[zip_index]
#                 zip_index += 1

#             # -----------------------------
#             # STORE RESULT
#             # -----------------------------
#             results.append({
#                 "unit": unit,
#                 "co": co,
#                 "unit_title": unit_title,
#                 "type": qtype,
#                 "part": part,
#                 "marks": marks,
#                 "bt": bt,
#                 "level": level,
#                 "questionText": question_text if question_text else None,
#                 "questionImage": question_image
#             })

#         return jsonify(results)

#     except Exception as e:
#         print("PYTHON EXTRACTION ERROR:\n", traceback.format_exc())
#         return jsonify({
#             "error": "Extraction failed",
#             "details": str(e)
#         }), 500

# # -----------------------------
# # Run app (local only)
# # -----------------------------
# if __name__ == "__main__":
#     import os
#     port = int(os.environ.get("PORT", 5001))
#     app.run(host="0.0.0.0", port=port)










from flask import Flask, request, jsonify
from openpyxl import load_workbook
import base64
import traceback
from flask_cors import CORS
import zipfile
from io import BytesIO
import xml.etree.ElementTree as ET

app = Flask(__name__)
CORS(app)

# -----------------------------
# Safe anchor reader (openpyxl)
# -----------------------------
def get_anchor(img):
    try:
        if hasattr(img.anchor, "_from"):
            return img.anchor._from.row + 1, img.anchor._from.col + 1
        elif hasattr(img.anchor, "from_"):
            return img.anchor.from_.row + 1, img.anchor.from_.col + 1
        else:
            return None, None
    except:
        return None, None

# -----------------------------
# Extract images with exact positions (XML method)
# -----------------------------
def extract_images_with_positions(file_bytes):
    image_map = {}

    try:
        with zipfile.ZipFile(BytesIO(file_bytes)) as z:

            drawing_files = [
                f for f in z.namelist()
                if f.startswith("xl/drawings/drawing") and f.endswith(".xml")
            ]

            for drawing in drawing_files:
                xml_data = z.read(drawing)
                tree = ET.fromstring(xml_data)

                ns = {
                    "xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing",
                    "a": "http://schemas.openxmlformats.org/drawingml/2006/main",
                    "r": "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
                }

                # Handle both anchor types
                anchors = tree.findall(".//xdr:oneCellAnchor", ns) + \
                          tree.findall(".//xdr:twoCellAnchor", ns)

                # Load relationships
                rels_path = drawing.replace("drawings/", "drawings/_rels/") + ".rels"
                rels_xml = z.read(rels_path)
                rels_tree = ET.fromstring(rels_xml)

                rel_dict = {
                    rel.attrib["Id"]: rel.attrib["Target"]
                    for rel in rels_tree
                }

                for anchor in anchors:
                    from_tag = anchor.find("xdr:from", ns)

                    row = int(from_tag.find("xdr:row", ns).text) + 1
                    col = int(from_tag.find("xdr:col", ns).text) + 1

                    blip = anchor.find(".//a:blip", ns)
                    if blip is not None:
                        embed = blip.attrib.get(
                            "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}embed"
                        )

                        if embed in rel_dict:
                            target = rel_dict[embed]
                            img_path = "xl/" + target.replace("../", "")

                            img_data = z.read(img_path)
                            base64_img = base64.b64encode(img_data).decode("utf-8")

                            image_map[(row, col)] = base64_img

    except Exception as e:
        print("XML mapping error:", str(e))

    return image_map

# -----------------------------
# Health check
# -----------------------------
@app.route("/")
def home():
    return "Flask API is running!"

# -----------------------------
# Main API
# -----------------------------
@app.route("/extract", methods=["POST"])
def extract_questions():
    try:
        if "file" not in request.files:
            return jsonify({"error": "No file uploaded"}), 400

        file = request.files["file"]

        file_bytes = file.read()
        file_stream = BytesIO(file_bytes)

        wb = load_workbook(file_stream, data_only=True)
        ws = wb.active

        results = []

        # openpyxl images
        images = getattr(ws, "_images", [])
        print("openpyxl images found:", len(images))

        # XML-based mapping
        image_map = extract_images_with_positions(file_bytes)
        print("XML mapped images:", len(image_map))

        # Loop rows
        for row_index, row in enumerate(ws.iter_rows(min_row=2), start=2):

            if not any(cell.value for cell in row):
                continue

            unit = row[0].value
            co = row[1].value
            unit_title = row[2].value
            qtype = row[3].value
            part = row[4].value
            question_text = row[5].value
            marks = row[6].value
            bt = row[7].value
            level = row[8].value

            if qtype:
                qtype = str(qtype).strip().upper()

            # -----------------------------
            # IMAGE DETECTION
            # -----------------------------
            question_image = None

            # 1️⃣ Try openpyxl (local)
            for img in images:
                try:
                    img_row, img_col = get_anchor(img)
                    if img_row == row_index and img_col == 6:
                        img_bytes = img._data()
                        question_image = base64.b64encode(img_bytes).decode("utf-8")
                        break
                except Exception as e:
                    print("Image error:", str(e))

            # 2️⃣ Fallback (XML mapping)
            if not question_image:
                question_image = image_map.get((row_index, 6))

            # -----------------------------
            # STORE RESULT
            # -----------------------------
            results.append({
                "unit": unit,
                "co": co,
                "unit_title": unit_title,
                "type": qtype,
                "part": part,
                "marks": marks,
                "bt": bt,
                "level": level,
                "questionText": question_text if question_text else None,
                "questionImage": question_image
            })

        return jsonify(results)

    except Exception as e:
        print("PYTHON EXTRACTION ERROR:\n", traceback.format_exc())
        return jsonify({
            "error": "Extraction failed",
            "details": str(e)
        }), 500

# -----------------------------
# Run app
# -----------------------------
if __name__ == "__main__":
    import os
    port = int(os.environ.get("PORT", 5001))
    app.run(host="0.0.0.0", port=port)